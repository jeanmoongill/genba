from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

import pprint
import json

class ProcessExcel:

    def main():

        CATEGORY = 1
        SHIHYO = 2
        KAMOKU = 3

        wb = load_workbook('./for_test.xlsx')
        ws = wb["Sheet1"]

        _temp_shihyo = None
        
        data_list = []
        shihyo_temp = None
        category_temp = None

        type_list = []
        _years_list = []
        _kizyun_list = []
        for row in range(1, ws.max_row):
            _value = ws.cell(row = row, column = 1).value

            type_dict = {}
            if _value in ['連携', '単体']:
                type_dict['type'] = _value
                type_dict['row'] = row
                type_list.append(type_dict)

        for index, row in enumerate(range(KAMOKU, ws.max_row + 1)):

            _data_dict = {}
            _shigyo = ws.cell(row = row, column = 2).value
            _kamoku = ws.cell(row = row, column = 3).value
            _category = ws.cell(row = row - 1, column = 1).value

            if _category is not None:
                category_temp = _category

            sub_dict = {}
            data = []

            _type = None
            for row_type in type_list:
                if row > row_type.get('row'):
                    _type = row_type.get('type')

            if _temp_shihyo != _shigyo and _shigyo is not None and _kamoku is not None:
                _kamoku_for_store = ws.cell(row = row - 1, column = KAMOKU).value
                if _kamoku_for_store:
                    
                    for data_row in range(1, 6):
                        _data = ws.cell(row = row - 1, column = KAMOKU + data_row).value
                        data.append(_data)
                
                    if _kamoku_for_store == '科目':
                        shihyo_temp = _shigyo
                        _years_list = data
                        continue
                    else:

                        sub_dict['type'] = _type
                        sub_dict['category'] = category_temp
                        sub_dict['shihyo'] = shihyo_temp
                        sub_dict['kamoku'] = _kamoku_for_store
                        sub_dict['data'] = data
                        sub_dict['years'] = _years_list

                        shihyo_temp = _shigyo
                    
                    data_list.append(sub_dict)
                    _temp_shihyo = _shigyo

            if _kamoku is None:

                _last_kamoku = ws.cell(row = row - 1, column = 3).value

                if _last_kamoku is not None:
                    last_komoku_list = []
                    _last_komoku_dict = {}
                    for _last_komoku in range(1, 6):
                        _last_value = ws.cell(row = row - 1, column = KAMOKU + _last_komoku).value
                        last_komoku_list.append(_last_value)

                    _last_komoku_dict['type'] = _type
                    _last_komoku_dict['category'] = category_temp
                    _last_komoku_dict['shihyo'] = shihyo_temp
                    _last_komoku_dict['kamoku'] = _last_kamoku
                    _last_komoku_dict['data'] = last_komoku_list   
                    _last_komoku_dict['years'] = _years_list
                    shihyo_temp = _shigyo

                    data_list.append(_last_komoku_dict)

                # 規準を格納する
                _kizyun_list = []
                _kizyun_dict = {}
                for _kizyun_col in range(1, 6):
                    _kizyun_data = ws.cell(row = row, column = KAMOKU + _kizyun_col).value
                    if _kizyun_data is not None:
                        _kizyun_list.append(_kizyun_data)

                if len(_kizyun_list) > 0:
                    for _get_data in data_list:
                        if 'kizyun' not in _get_data.keys():
                            _get_data['kizyun'] = _kizyun_list

        # draw chart
        _years_list = []
        _kizyun_list = []

        _data_data_list_result = []

        pprint.pprint(data_list)
        
        for _index_one, _dict_data in enumerate(data_list):
            _data_data_list = []
            _keys = list(_dict_data.keys()) 
           
            if 'data' in _keys and 'years' in _keys:
                pre_kizyun = None
                _diff_index = None
                for _index_two in range(0, len(_dict_data.get('years'))):
                    _tmp = []
                    _tmp.append(_dict_data.get('years')[_index_two])
                    
                    _tmp.append(_dict_data.get('type'))
                    _tmp.append(_dict_data.get('category'))
                    _tmp.append(_dict_data.get('shihyo'))
                    _tmp.append(_dict_data.get('kamoku'))
                    _tmp.append(_dict_data.get('kizyun')[_index_two])

                    _current_kizyun = _dict_data.get('kizyun')[_index_two]
                    _current_data = _dict_data.get('data')[_index_two]

                    if pre_kizyun != None and pre_kizyun != _current_kizyun or _current_data == '-':
                        # TODO: set datapoint
                        _tmp.append(_current_data)
                        _diff_index = _index_two

                    else:
                        _tmp.append(_current_data)

                    pre_kizyun = _current_kizyun

                    if len(_tmp) > 0:
                        _data_data_list.append(_tmp)

                # Draw data
                _data_data_list_result.append(_data_data_list)

        _chart_index = 0
        for index, data in enumerate(_data_data_list_result):

            ws = wb["Sheet2"]

            c1 = LineChart()
            c1.style = 13

            for _data in data:
                c1.title = _data[2] + "・" + _data[4]
                c1.x_axis.title = '年度'
                c1.y_axis.title = _data[3]
                c1.legend = None
                ws.append(_data)
            
            size_row = 1 + index * len(data)
            size_col = len(data)

            excel_data = Reference(
                ws,
                min_col = 7, 
                max_col = 7, 
                min_row = 1 + index * len(data), 
                max_row = (1 + index * len(data)) + len(data) - 1)

            c1.add_data(excel_data, titles_from_data=False)

            # style
            s1 = c1.series[0]
            s1.marker.graphicalProperties.solidFill = "FEFEFE"
            s1.marker.graphicalProperties.line.solidFill = "FEFEFE"
            s1.graphicalProperties.line.noFill = False

            pt = DataPoint(idx=3)
            pt.graphicalProperties.line.noFill = True
            s1.dPt.append(pt)

            ws.add_chart(c1, "I%s"%_chart_index + str(len(data)))
            wb.save("line2.xlsx")

            _chart_index += 2   
            # print('----------------------------')

    def output_chart(self, ws, data, row):
        rows = [
            ['Date', 'Batch 1'],
            [date(2015,9, 1), 40],
            [date(2015,9, 2), 40],
            [date(2015,9, 3)],
            [date(2015,9, 4), 30],
            [date(2015,9, 5), 25],
            [date(2015,9, 6), 20],
        ]

        for row in rows:
            ws.append(row)

        c1 = LineChart()
        c1.title = "test"
        c1.style = 13
        c1.y_axis.title = 'Y axis'
        c1.x_axis.title = 'X axis'

        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
        c1.add_data(data, titles_from_data=True)

        ws.add_chart(c1, "A10")
        wb.save("line.xlsx")
        pass

    def get_kizyun(self, row):
        """
        基準をセットする。
        """
        pass

    if __name__ == "__main__":
        main()